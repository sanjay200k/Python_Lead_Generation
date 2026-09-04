"""
gosom_lead_scraper.py
----------------------
Reads one row from your lead_search_criteria CSV, builds a gosom (free,
open-source, locally-run Google Maps scraper) query, runs it via Docker,
then filters the raw results against that row's criteria and writes:

  1) a per-row output, in TWO easy-to-read formats:
       - qualified_leads_row{N}.xlsx  (formatted spreadsheet -- open and go)
       - qualified_leads_row{N}.csv   (plain CSV, kept for compatibility)
  2) an upsert into a PERSISTENT master CSV that accumulates good leads
     across every row and every run you ever do (never wiped), plus a
     master_qualified_leads.xlsx snapshot regenerated from it every run
     so you always have a readable, formatted copy too.

CHANGE IN THIS VERSION: SQLite has been removed entirely. The master
store is now a single CSV (master_leads.csv) -- read with pandas,
deduped by `cid` (newest wins), rewritten each run. No Docker/DB-browser
dependency, no SQLite INTEGER overflow risk from large Google Maps `cid`
values (they're just written as plain text in a CSV, no type coercion
issue). If you ever want to query it, just `pd.read_csv(MASTER_CSV_PATH)`
and filter with pandas, or open it in Excel.

Requirements:
    - Docker Desktop installed and RUNNING
    - pip install pandas openpyxl

Usage:
    Set CSV_PATH and ROW_NUMBER below and run this file. Run it once per
    row (or loop over rows -- see run_all_rows() at the bottom).
"""

import os
import subprocess
from datetime import datetime, timezone

import pandas as pd

# ---------------- CONFIG ----------------
CSV_PATH = r"F:\AI automation\AI automation\Python_Lead_Generation - Copy\Python_Lead_Generation\Main_project\lead scaping data\lead_search_criteria (2).csv"
ROW_NUMBER = 3          # 1 = first row, 2 = second row, etc.
BASE_DEPTH = 5          # gosom scroll depth baseline; scaled by `priority` per row
EXIT_ON_INACTIVITY = "3m"
WORK_DIR = os.path.abspath("gosom_run")             # per-run raw scrape + per-row output files
DOCKER_IMAGE = "gosom/google-maps-scraper"
CACHE_VOLUME = "gmaps-playwright-cache"             # named volume (browser cache only)

# Persistent, NEVER wiped -- accumulates good leads across every row/run.
MASTER_CSV_PATH = os.path.abspath("master_leads.csv")
MASTER_EXCEL_PATH = os.path.abspath("master_qualified_leads.xlsx")

# Optional, user-maintained. Add a `cid` (preferred) or `business_name`
# column yourself after you've reached out to someone. If this file
# doesn't exist, "already contacted" exclusion is skipped with a note.
CONTACTED_LEADS_PATH = os.path.abspath("contacted_leads.csv")

PRIORITY_DEPTH_MAP = {"high": 7, "medium": 5, "low": 3}

LANGUAGE_TO_CODE = {
    "english": "en", "arabic": "ar", "french": "fr", "spanish": "es",
    "german": "de", "portuguese": "pt", "hindi": "hi", "tamil": "ta",
    "chinese": "zh", "japanese": "ja", "korean": "ko", "italian": "it",
    "russian": "ru", "dutch": "nl", "turkish": "tr",
}

# Candidate raw-column names gosom might use for each human-readable field.
OUTPUT_COLUMN_CANDIDATES = {
    "business name": ["title", "name", "business_name"],
    "category": ["category", "categories"],
    "address": ["address", "complete_address", "full_address"],
    "phone": ["phone", "phone_number"],
    "website": ["website", "site"],
    "email": ["emails", "email"],
    "google maps url": ["link", "google_maps_link", "url"],
    "rating": ["review_rating", "rating"],
    "review count": ["review_count", "reviews"],
    "social media urls": ["social_media", "social_links", "socials"],
}
URL_LABELS = {"website", "google maps url"}  # get hyperlinked in the Excel output
NAME_LOOKUP_CANDIDATES = ["title", "name", "business_name"]
CATEGORY_LOOKUP_CANDIDATES = ["category", "categories"]
# -----------------------------------------


def load_criteria_row(csv_path, row_number):
    df = pd.read_csv(csv_path)
    df = df.dropna(how="all").reset_index(drop=True)
    total_rows = len(df)
    if row_number < 1 or row_number > total_rows:
        raise ValueError(f"Out of range! CSV has {total_rows} rows, but you asked for row {row_number}.")
    return df.iloc[row_number - 1]


def build_location(row):
    parts = []
    for col in ("city_area", "area_extra", "state_province", "country"):
        val = row.get(col)
        if pd.notna(val) and str(val).strip():
            parts.append(str(val).strip())
    return ", ".join(parts)


def build_queries(row):
    location = build_location(row)
    terms = set()

    business_category = row.get("business_category")
    if pd.notna(business_category) and str(business_category).strip():
        terms.add(str(business_category).strip())

    keywords = row.get("keywords")
    if pd.notna(keywords) and str(keywords).strip():
        for kw in str(keywords).split(";"):
            kw = kw.strip()
            if kw:
                terms.add(kw)

    if not terms:
        print("Warning: row has no business_category or keywords -- no queries to run.")
        return []

    return [f"{term} in {location}" for term in terms]


def resolve_column(df, candidates):
    for c in candidates:
        if c in df.columns:
            return c
    return None


def resolve_depth_and_lang(row):
    priority = str(row.get("priority", "")).strip().lower()
    depth = PRIORITY_DEPTH_MAP.get(priority, BASE_DEPTH)

    lang_code = None
    language_raw = row.get("language")
    if pd.notna(language_raw) and str(language_raw).strip():
        first_lang = str(language_raw).split(";")[0].strip().lower()
        lang_code = LANGUAGE_TO_CODE.get(first_lang)
        if lang_code is None:
            print(f"Note: language '{first_lang}' has no known gosom locale code -- "
                  f"leaving -lang unset for this run.")
    return depth, lang_code


def run_gosom(queries, work_dir, row_number, depth, lang_code):
    os.makedirs(work_dir, exist_ok=True)
    queries_path = os.path.join(work_dir, "queries.txt")
    results_path = os.path.join(work_dir, f"raw_results_row{row_number}.csv")

    with open(queries_path, "w", encoding="utf-8") as f:
        f.write("\n".join(queries))
    open(results_path, "w").close()

    cmd = [
        "docker", "run", "--rm",
        "-v", f"{CACHE_VOLUME}:/opt",
        "-v", f"{queries_path}:/queries.txt:ro",
        "-v", f"{results_path}:/results.csv",
        DOCKER_IMAGE,
        "-input", "/queries.txt",
        "-results", "/results.csv",
        "-depth", str(depth),
        "-exit-on-inactivity", EXIT_ON_INACTIVITY,
    ]
    if lang_code:
        cmd += ["-lang", lang_code]

    print("Running gosom via Docker (first run may take a while to pull the image)...")
    print(" ".join(cmd))
    subprocess.run(cmd, check=True)
    return results_path


def normalize_raw_csv(results_path):
    """
    gosom is expected to write one row per business. But raw_results_row1.csv
    (a sample of this pipeline's actual output) turned out to be TRANSPOSED
    instead: a 'Column' header, field names running down the second column
    (C1=input_id, C2=link, ...), and each business's values spread across
    dozens of repeated columns to the right -- array fields like
    user_reviews/about forced every scalar field to repeat once per array
    element, so 4 real businesses turned into 144 value-columns.

    Reading that shape with plain pd.read_csv() gives you fields-as-rows and
    businesses-as-columns, which breaks every filter below. This detects
    that shape and reshapes it back to one row per business first. If your
    gosom output is already normal (one row per business), this is a no-op.
    """
    df_raw = pd.read_csv(results_path)
    if df_raw.empty:
        return df_raw

    first_col = df_raw.columns[0]
    looks_transposed = (
        str(first_col).strip().lower() == "column"
        and df_raw.shape[1] > 1
        and not bool((df_raw.iloc[:, 0].astype(str) == "input_id").any())  # field names live in col 1, not col 0
        and bool((df_raw.iloc[:, 0].astype(str).str.match(r"^C\d+$")).all())
    )
    if not looks_transposed:
        return df_raw

    print(f"Note: {results_path} is in transposed format (fields running down "
          f"rows instead of across columns) -- reshaping to one row per "
          f"business before filtering.")

    field_names = df_raw.iloc[:, 1].tolist()
    value_cols = list(df_raw.columns[2:])

    if "input_id" not in field_names:
        print("Warning: transposed raw results have no 'input_id' row -- "
              "can't tell where one business ends and the next begins. "
              "Returning as-is; filtering will likely fail.")
        return df_raw

    input_id_row_idx = field_names.index("input_id")
    input_id_values = df_raw.iloc[input_id_row_idx, 2:].tolist()

    records = []
    start = 0
    for i in range(1, len(input_id_values) + 1):
        if i == len(input_id_values) or input_id_values[i] != input_id_values[start]:
            col = value_cols[start]  # first column belonging to this business
            record = {field_names[r]: df_raw.iloc[r][col] for r in range(len(field_names))}
            records.append(record)
            start = i

    print(f"Reshaped {len(input_id_values)} value-columns into {len(records)} business rows.")
    return pd.DataFrame(records)


def apply_contact_requirements(df, row):
    def has_value(col):
        return df[col].notna() & (df[col].astype(str).str.strip() != "")

    mask = pd.Series(True, index=df.index)

    # Website: true tri-state. "No" is a legitimate targeting signal
    # (businesses WITHOUT a website -- e.g. if you're pitching website
    # or automation services to them).
    website_required = str(row.get("website_required", "")).strip().lower()
    if website_required == "yes" and "website" in df.columns:
        mask &= has_value("website")
    elif website_required == "no" and "website" in df.columns:
        mask &= ~has_value("website")
    # "either"/blank -> no filter

    # Phone & email: "Yes" filters for must-have. Anything else means
    # "not required" -- NEVER interpreted as "must not have", because
    # you never want to throw away a lead specifically for having
    # contact info you could use.
    phone_required = str(row.get("phone_required", "")).strip().lower()
    if phone_required == "yes" and "phone" in df.columns:
        mask &= has_value("phone")

    email_required = str(row.get("email_required", "")).strip().lower()
    email_col = resolve_column(df, ["emails", "email"])
    if email_required == "yes" and email_col:
        mask &= has_value(email_col)

    return mask


def apply_exclude_terms(filtered, row, master_history_cids, contacted_ids):
    exclude_terms_raw = row.get("exclude_terms")
    if not (pd.notna(exclude_terms_raw) and str(exclude_terms_raw).strip()):
        return filtered

    raw_terms = [t.strip().lower() for t in str(exclude_terms_raw).split(";") if t.strip()]
    special_terms = {"duplicates", "already contacted"}
    literal_terms = [t for t in raw_terms if t not in special_terms]

    before = len(filtered)

    # Literal substring terms against name/category
    if literal_terms:
        name_col = resolve_column(filtered, NAME_LOOKUP_CANDIDATES)
        cat_col = resolve_column(filtered, CATEGORY_LOOKUP_CANDIDATES)
        if name_col or cat_col:
            def is_excluded(r):
                haystack = ""
                if name_col:
                    haystack += str(r.get(name_col, "")).lower() + " "
                if cat_col:
                    haystack += str(r.get(cat_col, "")).lower()
                return any(term in haystack for term in literal_terms)
            filtered = filtered[~filtered.apply(is_excluded, axis=1)]
        else:
            print(f"Warning: exclude_terms has literal terms {literal_terms} but no "
                  f"business-name/category column was found -- skipped.")
        print(f"exclude_terms (literal: {', '.join(literal_terms)}) removed "
              f"{before - len(filtered)} lead(s).")
        before = len(filtered)

    # "duplicates" -> drop anything already in the persistent master
    if "duplicates" in raw_terms:
        if "cid" in filtered.columns:
            filtered = filtered[~filtered["cid"].astype(str).isin(master_history_cids)]
            print(f"exclude_terms 'duplicates' removed {before - len(filtered)} "
                  f"lead(s) already present in {MASTER_CSV_PATH}.")
        else:
            print("Warning: exclude_terms has 'duplicates' but no 'cid' column "
                  "exists in raw results -- skipped.")
        before = len(filtered)

    # "already contacted" -> drop anything in contacted_leads.csv
    if "already contacted" in raw_terms:
        if contacted_ids is None:
            print(f"Note: exclude_terms has 'already contacted' but "
                  f"{CONTACTED_LEADS_PATH} doesn't exist yet -- skipped. "
                  f"Create it with a 'cid' or 'business_name' column to enable this.")
        elif "cid" in filtered.columns:
            filtered = filtered[~filtered["cid"].astype(str).isin(contacted_ids)]
            print(f"exclude_terms 'already contacted' removed "
                  f"{before - len(filtered)} lead(s).")

    return filtered


def load_master_csv(csv_path=MASTER_CSV_PATH):
    """Load the persistent master CSV. Returns an empty DataFrame if it
    doesn't exist yet or is empty/corrupt (never raises)."""
    if not os.path.exists(csv_path):
        return pd.DataFrame()
    try:
        return pd.read_csv(csv_path)
    except pd.errors.EmptyDataError:
        return pd.DataFrame()


def load_master_history_cids(csv_path=MASTER_CSV_PATH):
    existing = load_master_csv(csv_path)
    if "cid" in existing.columns:
        return set(existing["cid"].dropna().astype(str))
    return set()


def load_contacted_ids(contacted_path):
    if not os.path.exists(contacted_path):
        return None
    try:
        df = pd.read_csv(contacted_path)
    except pd.errors.EmptyDataError:
        return set()
    if "cid" in df.columns:
        return set(df["cid"].dropna().astype(str))
    if "business_name" in df.columns:
        return set(df["business_name"].dropna().astype(str).str.lower())
    return set()


def safe_float(value, default=None):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def safe_int(value, default=None):
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def filter_leads(results_path, row, master_history_cids, contacted_ids):
    try:
        df = normalize_raw_csv(results_path)
    except pd.errors.EmptyDataError:
        print("Warning: gosom produced no data for this run (empty raw results).")
        return pd.DataFrame()
    if df.empty:
        return df

    min_rating = safe_float(row.get("min_rating"), default=0.0)
    min_reviews = safe_int(row.get("min_reviews"), default=0)
    if min_rating is None:
        print(f"Warning: min_rating '{row.get('min_rating')}' isn't a valid number -- treating as no filter (0).")
        min_rating = 0.0
    if min_reviews is None:
        print(f"Warning: min_reviews '{row.get('min_reviews')}' isn't a valid number -- treating as no filter (0).")
        min_reviews = 0

    df["review_rating"] = pd.to_numeric(df.get("review_rating"), errors="coerce")
    df["review_count"] = pd.to_numeric(df.get("review_count"), errors="coerce")

    mask = (df["review_rating"] >= min_rating) & (df["review_count"] >= min_reviews)
    mask &= apply_contact_requirements(df, row)

    filtered = df[mask].copy()

    # Dedupe within this run's results
    dedupe_col = "cid" if "cid" in filtered.columns else "link"
    if dedupe_col in filtered.columns:
        filtered = filtered.drop_duplicates(subset=[dedupe_col])

    filtered = apply_exclude_terms(filtered, row, master_history_cids, contacted_ids)

    # search_radius_km: gosom takes a text query, not lat/long + radius --
    # there's no geocoding step in this script to enforce a radius, so this
    # is surfaced rather than silently pretending it narrowed the search.
    radius = row.get("search_radius_km")
    if pd.notna(radius) and str(radius).strip():
        print(f"Note: search_radius_km = '{radius}' is NOT enforced -- gosom "
              f"searches by text query only, this script has no geocoding "
              f"step to filter by distance.")

    # review_quality_filter: format is inconsistent across rows in the
    # current CSV (e.g. "Outdated; average; good; none" isn't a single
    # value) -- surfaced, not enforced, so you know it's being skipped.
    review_quality = row.get("review_quality_filter")
    if pd.notna(review_quality) and str(review_quality).strip().lower() not in ("", "any"):
        print(f"Note: review_quality_filter = '{review_quality}' is not implemented "
              f"(inconsistent format across rows) -- IGNORED this run.")

    # toggle_1..5: undefined meaning. Print raw values so nothing is
    # silently dropped without visibility.
    toggle_values = {c: row.get(c) for c in ("toggle_1", "toggle_2", "toggle_3", "toggle_4", "toggle_5")
                      if c in row.index}
    if toggle_values:
        print(f"Note: toggle columns for this row (meaning not yet defined, "
              f"not applied): {toggle_values}")

    # Sort BEFORE capping so max_leads keeps your best candidates, not
    # just whichever ones happened to scrape first.
    sort_cols = [c for c in ("review_count", "review_rating") if c in filtered.columns]
    if sort_cols:
        filtered = filtered.sort_values(by=sort_cols, ascending=False)

    qualified_before_cap = len(filtered)
    max_leads = safe_int(row.get("max_leads"), default=None)
    if pd.notna(row.get("max_leads")) and max_leads is None:
        print(f"Warning: max_leads value '{row.get('max_leads')}' isn't valid -- "
              f"ignoring cap, returning all qualified leads.")

    if max_leads is not None:
        filtered = filtered.head(max_leads)
    else:
        print("Note: max_leads is blank/invalid -- no cap applied.")

    print(f"Requested max_leads: {max_leads if max_leads is not None else 'none set'} "
          f"| Qualified before cap: {qualified_before_cap} | Delivered: {len(filtered)}")

    return filtered


def reshape_output_columns(filtered, row):
    """Build the DISPLAY dataframe per output_columns. The caller keeps the
    original `filtered` (with cid etc.) separately for the master upsert."""
    output_columns_raw = row.get("output_columns")
    if not (pd.notna(output_columns_raw) and str(output_columns_raw).strip()):
        return filtered.copy()

    requested_labels = [c.strip() for c in str(output_columns_raw).split(";") if c.strip()]
    selected = {}
    for label in requested_labels:
        candidates = OUTPUT_COLUMN_CANDIDATES.get(label.strip().lower())
        actual_col = resolve_column(filtered, candidates) if candidates else None
        if actual_col:
            selected[label] = filtered[actual_col]
        else:
            print(f"Warning: output_columns requested '{label}' but no matching "
                  f"column was found -- left blank.")
            selected[label] = pd.Series([""] * len(filtered), index=filtered.index)
    return pd.DataFrame(selected)


def write_excel(df, path, url_labels=None):
    """Write a DataFrame to a formatted, easy-to-read .xlsx: bold frozen
    header row, autofit column widths, autofilter, and clickable hyperlinks
    on any column named in url_labels."""
    if df.empty:
        print(f"Note: nothing to write to {path} (0 rows).")
        return

    url_labels = url_labels or set()
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Leads")
        ws = writer.sheets["Leads"]

        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        header_font = Font(bold=True)
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font

            # Autofit (roughly) based on the widest value in the column
            max_len = max(
                [len(str(col_name))] + [len(str(v)) for v in df.iloc[:, col_idx - 1].tolist()]
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

            if str(col_name).strip().lower() in url_labels:
                for row_idx in range(2, len(df) + 2):
                    c = ws.cell(row=row_idx, column=col_idx)
                    if c.value and str(c.value).startswith("http"):
                        c.hyperlink = str(c.value)
                        c.style = "Hyperlink"

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    print(f"Wrote {len(df)} row(s) to {path}")


def upsert_master_csv(filtered, row, row_number, csv_path=MASTER_CSV_PATH):
    """Add/refresh this run's qualified leads into the persistent master
    CSV, tagged with source info, deduped by cid (newest data wins)."""
    if filtered.empty:
        return load_master_csv(csv_path)

    tagged = filtered.copy()
    tagged.insert(0, "source_row", row_number)
    tagged.insert(1, "source_business_category", row.get("business_category"))
    tagged.insert(2, "source_location", build_location(row))
    tagged["added_at"] = datetime.now(timezone.utc).isoformat()

    existing = load_master_csv(csv_path)
    dedupe_col = "cid" if "cid" in tagged.columns else "link"
    combined = pd.concat([existing, tagged], ignore_index=True, sort=False)
    if dedupe_col in combined.columns:
        # Keep cid/link as plain strings -- these are opaque IDs, not
        # numbers to do math on, and large Google Maps cid values can be
        # far bigger than a normal int, so writing them as text avoids
        # any type-conversion surprises when this CSV is reopened later.
        combined[dedupe_col] = combined[dedupe_col].astype(str)
        combined = combined.drop_duplicates(subset=[dedupe_col], keep="last")

    combined.to_csv(csv_path, index=False)
    print(f"Master CSV now has {len(combined)} total accumulated leads: {csv_path}")
    return combined


def run_one_row(row_number):
    row = load_criteria_row(CSV_PATH, row_number)
    print(f"\n=== Row {row_number}: {row['business_category']} in {build_location(row)} ===")

    depth, lang_code = resolve_depth_and_lang(row)
    queries = build_queries(row)
    print("Queries to run:")
    for q in queries:
        print(" -", q)
    if not queries:
        print("Nothing to run for this row -- skipping.")
        return

    raw_results_path = run_gosom(queries, WORK_DIR, row_number, depth, lang_code)

    master_history_cids = load_master_history_cids(MASTER_CSV_PATH)
    contacted_ids = load_contacted_ids(CONTACTED_LEADS_PATH)

    qualified = filter_leads(raw_results_path, row, master_history_cids, contacted_ids)

    # Master gets the full-detail upsert (needed for future dedupe/history)
    master_df = upsert_master_csv(qualified, row, row_number, MASTER_CSV_PATH)
    write_excel(master_df, MASTER_EXCEL_PATH, url_labels={"website", "link"})

    # Per-row output: reshaped, human-facing, in both xlsx and csv
    display_df = reshape_output_columns(qualified, row)
    row_csv_path = os.path.join(WORK_DIR, f"qualified_leads_row{row_number}.csv")
    row_xlsx_path = os.path.join(WORK_DIR, f"qualified_leads_row{row_number}.xlsx")
    display_df.to_csv(row_csv_path, index=False)
    write_excel(display_df, row_xlsx_path, url_labels=URL_LABELS)

    print(f"Raw results: {raw_results_path}")
    print(f"Row output ({len(display_df)} leads): {row_xlsx_path}  /  {row_csv_path}")


def run_all_rows():
    """Loop over every row in the criteria CSV, one gosom run per row."""
    df = pd.read_csv(CSV_PATH).dropna(how="all").reset_index(drop=True)
    for i in range(1, len(df) + 1):
        run_one_row(i)


def main():
    run_one_row(ROW_NUMBER)
    # To run every row in the CSV instead, comment the line above and
    # uncomment the line below:
    # run_all_rows()


if __name__ == "__main__":
    main()